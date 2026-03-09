"""
validator.py - Validator kapacitního Excelu
Kontroluje Excel soubory kapacit dle definovaných pravidel.
"""

import tkinter as tk
from tkinter import filedialog
from dataclasses import dataclass, field
from datetime import datetime
import os
import sys

# ── Volitelná závislost Pillow (pro zobrazení loga) ──────────────────────────
try:
    from PIL import Image, ImageTk
    _PILLOW_OK = True
except ImportError:
    _PILLOW_OK = False

# ── Kontrola závislosti openpyxl ─────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    import subprocess
    root = tk.Tk()
    root.withdraw()
    from tkinter import messagebox
    answer = messagebox.askyesno(
        "Chybí závislost",
        "Knihovna 'openpyxl' není nainstalována.\n\n"
        "Chcete ji automaticky nainstalovat?"
    )
    root.destroy()
    if answer:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
    else:
        sys.exit(1)

from tkinter import messagebox

# ── Konstanty ────────────────────────────────────────────────────────────────
CHECK_START_ROW = 14
CONSECUTIVE_EMPTY_STOP = 5
COL_B_INDEX = 2
CHECKED_COL_INDICES = [1, 3, 4, 5, 6, 7, 8]   # A, C, D, E, F, G, H
CHECKED_COL_LETTERS = ["A", "C", "D", "E", "F", "G", "H"]


# ── Česká abeceda – řadicí klíč ───────────────────────────────────────────────
_CZ_ALPHA = [
    'a', 'á', 'b', 'c', 'č', 'd', 'ď', 'e', 'é', 'ě',
    'f', 'g', 'h', 'ch', 'i', 'í', 'j', 'k', 'l', 'm',
    'n', 'ň', 'o', 'ó', 'p', 'q', 'r', 'ř', 's', 'š',
    't', 'ť', 'u', 'ú', 'ů', 'v', 'w', 'x', 'y', 'ý',
    'z', 'ž',
]
_CZ_ORDER = {ch: i for i, ch in enumerate(_CZ_ALPHA)}


def czech_sort_key(name: str) -> tuple:
    """Vrátí řadicí klíč pro českou abecedu; správně řeší digraf 'ch'."""
    s = name.lower()
    tokens = []
    i = 0
    while i < len(s):
        if s[i:i + 2] == 'ch':
            tokens.append('ch')
            i += 2
        else:
            tokens.append(s[i])
            i += 1
    return tuple(_CZ_ORDER.get(t, len(_CZ_ALPHA) + ord(t[0])) for t in tokens)


# ── Datové třídy ─────────────────────────────────────────────────────────────
@dataclass
class CellError:
    column_letter: str
    error_type: str       # 'empty' | 'excel_error'
    raw_value: str


@dataclass
class RowError:
    sheet_name: str
    row_number: int
    project_code: str
    cell_errors: list = field(default_factory=list)


@dataclass
class SheetResult:
    sheet_name: str
    checked: bool
    skip_reason: str = ""
    row_errors: list = field(default_factory=list)
    rows_checked: int = 0


@dataclass
class ValidationResult:
    file_path: str
    sheet_results: list = field(default_factory=list)
    total_errors: int = 0
    timestamp: str = ""
    sheet_order_violations: list = field(default_factory=list)
    # Každá položka: (název_listu, aktuální_pozice_1, správná_pozice_1)


# ── Logika validace ───────────────────────────────────────────────────────────
class ExcelValidator:

    def validate_file(self, file_path: str) -> ValidationResult:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_results = []

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("_"):
                sheet_results.append(SheetResult(
                    sheet_name=sheet_name,
                    checked=False,
                    skip_reason="Název listu začíná '_'",
                ))
            else:
                ws = wb[sheet_name]
                sheet_results.append(self._validate_sheet(sheet_name, ws))

        wb.close()

        total_errors = sum(
            len(sr.row_errors) for sr in sheet_results if sr.checked
        )

        # Kontrola pořadí listů dle české abecedy
        checked_names = [sr.sheet_name for sr in sheet_results if sr.checked]
        expected_order = sorted(checked_names, key=czech_sort_key)
        sheet_order_violations = [
            (name, checked_names.index(name) + 1, expected_order.index(name) + 1)
            for name in checked_names
            if checked_names.index(name) != expected_order.index(name)
        ]

        return ValidationResult(
            file_path=file_path,
            sheet_results=sheet_results,
            total_errors=total_errors,
            timestamp=datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
            sheet_order_violations=sheet_order_violations,
        )

    def _validate_sheet(self, sheet_name: str, ws) -> SheetResult:
        row_errors = []
        rows_checked = 0
        consecutive_empty = 0
        row = CHECK_START_ROW

        while True:
            b_val = ws.cell(row=row, column=COL_B_INDEX).value
            b_empty = b_val is None or str(b_val).strip() == ""

            if b_empty:
                consecutive_empty += 1
                if consecutive_empty >= CONSECUTIVE_EMPTY_STOP:
                    break
                row += 1
                continue

            consecutive_empty = 0
            rows_checked += 1
            project_code = str(b_val).strip()

            cell_errors = self._check_row(ws, row)
            if cell_errors:
                row_errors.append(RowError(
                    sheet_name=sheet_name,
                    row_number=row,
                    project_code=project_code,
                    cell_errors=cell_errors,
                ))
            row += 1

        return SheetResult(
            sheet_name=sheet_name,
            checked=True,
            row_errors=row_errors,
            rows_checked=rows_checked,
        )

    def _check_row(self, ws, row: int) -> list:
        errors = []
        for col_idx, col_letter in zip(CHECKED_COL_INDICES, CHECKED_COL_LETTERS):
            val = ws.cell(row=row, column=col_idx).value
            error_type = self._classify_cell(val)
            if error_type:
                errors.append(CellError(
                    column_letter=col_letter,
                    error_type=error_type,
                    raw_value=str(val) if val is not None else "",
                ))
        return errors

    def _classify_cell(self, value):
        if value is None:
            return "empty"
        s = str(value).strip()
        if s == "":
            return "empty"
        if s.startswith("#"):
            return "excel_error"
        return None


# ── Formátování reportu ───────────────────────────────────────────────────────
class ReportFormatter:

    ERROR_LABELS = {
        "empty": "Prázdné buňky",
        "excel_error": "Chybová hodnota Excel (#N/A)",
    }

    def _group_errors(self, cell_errors):
        """Seskupí chyby podle typu; vrátí seznam (popis, 'sl. A, C, D...')."""
        groups = {}
        for ce in cell_errors:
            groups.setdefault(ce.error_type, []).append(ce.column_letter)
        result = []
        for error_type, cols in groups.items():
            label = self.ERROR_LABELS.get(error_type, error_type)
            cols_str = "sl. " + ", ".join(cols)
            result.append((label, cols_str))
        return result

    def format_plain(self, result: ValidationResult) -> str:
        """Plain text pro uložení do souboru."""
        lines = []
        lines.append("=" * 64)
        lines.append("REPORT VALIDACE KAPACIT")
        lines.append(f"Soubor: {os.path.basename(result.file_path)}")
        lines.append(f"Cesta:  {result.file_path}")
        lines.append(f"Datum:  {result.timestamp}")
        lines.append("=" * 64)
        lines.append("")

        checked = [s for s in result.sheet_results if s.checked]
        skipped = [s for s in result.sheet_results if not s.checked]

        lines.append(
            f"Zkontrolované listy ({len(checked)}): "
            + ", ".join(s.sheet_name for s in checked)
        )
        if skipped:
            lines.append(
                f"Přeskočené listy ({len(skipped)}): "
                + ", ".join(s.sheet_name for s in skipped)
            )
        lines.append("")

        if not result.sheet_order_violations:
            lines.append("POŘADÍ LISTŮ: OK")
        else:
            lines.append("POŘADÍ LISTŮ: CHYBA")
            lines.append("  Listy na špatné pozici:")
            for name, actual, expected in result.sheet_order_violations:
                lines.append(
                    f"    * {name}  (aktuálně pozice {actual}, správně pozice {expected})"
                )
        lines.append("")

        for sr in result.sheet_results:
            if not sr.checked:
                continue
            lines.append(f"=== List: {sr.sheet_name} ===")
            if not sr.row_errors:
                lines.append("  OK - žádné chyby nenalezeny")
            else:
                for re in sr.row_errors:
                    for label, cols_str in self._group_errors(re.cell_errors):
                        lines.append(
                            f"  [CHYBA] Řádek {re.row_number}"
                            f" | Projekt: {re.project_code}"
                            f" | {label} ({cols_str})"
                        )
            lines.append("")

        sheets_with_errors = sum(1 for s in checked if s.row_errors)
        lines.append("-" * 64)
        lines.append(
            f"Celkem problematických řádků: {result.total_errors} | "
            f"Listy s chybami: {sheets_with_errors} z {len(checked)}"
        )
        lines.append("=" * 64)
        return "\n".join(lines)

    def segments_for_display(self, result: ValidationResult):
        """
        Vrátí seznam (text, tag) pro barevný výpis v tk.Text widgetu.
        Tagy: 'header', 'meta', 'sheet_ok', 'sheet_err', 'row_err',
              'cell_err', 'summary', 'normal'
        """
        segs = []

        def add(text, tag="normal"):
            segs.append((text, tag))

        add("=" * 64 + "\n", "header")
        add("REPORT VALIDACE KAPACIT\n", "header")
        add(f"Soubor: {os.path.basename(result.file_path)}\n", "meta")
        add(f"Cesta:  {result.file_path}\n", "meta")
        add(f"Datum:  {result.timestamp}\n", "meta")
        add("=" * 64 + "\n\n", "header")

        checked = [s for s in result.sheet_results if s.checked]
        skipped = [s for s in result.sheet_results if not s.checked]

        add(
            f"Zkontrolované listy ({len(checked)}): "
            + ", ".join(s.sheet_name for s in checked) + "\n",
            "meta",
        )
        if skipped:
            add(
                f"Přeskočené listy ({len(skipped)}): "
                + ", ".join(s.sheet_name for s in skipped) + "\n",
                "skipped",
            )
        add("\n")

        add("POŘADÍ LISTŮ: ", "sheet_header")
        if not result.sheet_order_violations:
            add("OK\n", "sheet_ok")
        else:
            add("CHYBA\n", "order_err")
            add("  Listy na špatné pozici:\n", "order_err")
            for name, actual, expected in result.sheet_order_violations:
                add(
                    f"    \u2022 {name}  "
                    f"(aktu\u00e1ln\u011b pozice {actual}, spr\u00e1vn\u011b pozice {expected})\n",
                    "order_err",
                )
        add("\n")

        for sr in result.sheet_results:
            if not sr.checked:
                continue
            add(f"=== List: {sr.sheet_name} ===\n", "sheet_header")
            if not sr.row_errors:
                add("  OK – žádné chyby nenalezeny\n", "sheet_ok")
            else:
                for re in sr.row_errors:
                    for label, cols_str in self._group_errors(re.cell_errors):
                        add(f"  [CHYBA] Řádek {re.row_number}", "row_err")
                        add(f" | Projekt: {re.project_code}", "row_err")
                        add(f" | ", "cell_err")
                        add(f"{label}", "cell_err")
                        add(f" ({cols_str})\n", "cell_err")
            add("\n")

        sheets_with_errors = sum(1 for s in checked if s.row_errors)
        add("-" * 64 + "\n", "summary")
        add(
            f"Celkem problematických řádků: {result.total_errors} | "
            f"Listy s chybami: {sheets_with_errors} z {len(checked)}\n",
            "summary",
        )
        add("=" * 64 + "\n", "header")
        return segs


# ── GUI aplikace ──────────────────────────────────────────────────────────────
class ValidatorApp(tk.Tk):

    # ── Pomocná metoda pro cesty assetů ──────────────────────────────────────
    @staticmethod
    def _asset_path(filename: str) -> str:
        """Vrátí absolutní cestu k souboru v graphics/; funguje i v .exe."""
        if getattr(sys, "frozen", False):
            base = sys._MEIPASS  # type: ignore[attr-defined]
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, "graphics", filename)

    def __init__(self):
        super().__init__()
        self.title("Validator kapacit v1.1")
        self.geometry("800x600")
        self.minsize(640, 480)
        self.resizable(True, True)

        self.validator = ExcelValidator()
        self.formatter = ReportFormatter()
        self.last_result = None
        self._logo_photo = None   # uchování reference (GC ochrana)
        self._icon_photo = None

        self._load_brand_images()
        self._build_ui()
        self._configure_tags()

    def _load_brand_images(self):
        """Načte PNG assety přes Pillow; při chybě nastaví None (graceful fallback)."""
        if not _PILLOW_OK:
            return

        # Symbol do headeru (40×40 px pro ostrost na HiDPI zobrazíme 40 px)
        try:
            sym_path = self._asset_path("parama-symbol.png")
            img = Image.open(sym_path).resize((40, 40), Image.LANCZOS)
            self._logo_photo = ImageTk.PhotoImage(img)
        except Exception:
            self._logo_photo = None

        # Symbol jako ikona titulního pruhu okna (32×32)
        try:
            icon_path = self._asset_path("parama-icon.png")
            ico = Image.open(icon_path).resize((32, 32), Image.LANCZOS)
            self._icon_photo = ImageTk.PhotoImage(ico)
        except Exception:
            self._icon_photo = None

    def _build_ui(self):
        # ── Ikona okna ───────────────────────────────────────────────────────
        if self._icon_photo:
            self.iconphoto(True, self._icon_photo)

        # ── Firemní hlavička ─────────────────────────────────────────────────
        hdr = tk.Frame(self, bg="#0D1B2A", padx=12, pady=8)
        hdr.pack(fill=tk.X)

        if self._logo_photo:
            tk.Label(
                hdr, image=self._logo_photo,
                bg="#0D1B2A", bd=0,
            ).pack(side=tk.LEFT, padx=(0, 10))

        txt_box = tk.Frame(hdr, bg="#0D1B2A")
        txt_box.pack(side=tk.LEFT)
        tk.Label(
            txt_box, text="PARAMA Software",
            font=("Segoe UI", 11, "bold"),
            bg="#0D1B2A", fg="#F4F6F9",
        ).pack(anchor=tk.W)
        tk.Label(
            txt_box, text="Validator kapacit",
            font=("Segoe UI", 8),
            bg="#0D1B2A", fg="#00C2A8",
        ).pack(anchor=tk.W)

        # ── Výběr souboru ────────────────────────────────────────────────────
        top = tk.Frame(self, padx=12, pady=10)
        top.pack(fill=tk.X)

        tk.Label(top, text="Soubor Excel:", font=("Segoe UI", 9)).grid(
            row=0, column=0, sticky=tk.W, pady=(0, 2)
        )

        self.path_var = tk.StringVar()
        self.path_entry = tk.Entry(
            top, textvariable=self.path_var,
            state="readonly", font=("Segoe UI", 9), relief=tk.SUNKEN, bd=1
        )
        self.path_entry.grid(row=1, column=0, sticky=tk.EW, ipady=3)

        tk.Button(
            top, text="Vybrat…", command=self._pick_file,
            font=("Segoe UI", 9), padx=10
        ).grid(row=1, column=1, padx=(6, 0))

        top.columnconfigure(0, weight=1)

        # ── Tlačítka ─────────────────────────────────────────────────────────
        btn_frame = tk.Frame(self, padx=12, pady=4)
        btn_frame.pack(fill=tk.X)

        self.run_btn = tk.Button(
            btn_frame,
            text="▶  Spustit kontrolu",
            command=self._run_validation,
            state=tk.DISABLED,
            font=("Segoe UI", 10, "bold"),
            bg="#0078D4", fg="white",
            activebackground="#005A9E", activeforeground="white",
            relief=tk.FLAT, bd=0, padx=18, pady=6, cursor="hand2",
        )
        self.run_btn.pack(side=tk.LEFT)

        self.save_btn = tk.Button(
            btn_frame,
            text="💾  Uložit report (.txt)",
            command=self._save_report,
            state=tk.DISABLED,
            font=("Segoe UI", 9),
            padx=12, pady=6, relief=tk.GROOVE,
        )
        self.save_btn.pack(side=tk.LEFT, padx=(8, 0))

        # ── Výsledky ──────────────────────────────────────────────────────────
        lbl_frame = tk.Frame(self)
        lbl_frame.pack(fill=tk.X, padx=12, pady=(4, 0))
        tk.Label(
            lbl_frame, text="Výsledky kontroly",
            font=("Segoe UI", 9, "bold"), fg="#333333"
        ).pack(side=tk.LEFT)

        text_frame = tk.Frame(self)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=(4, 0))

        scrollbar_y = tk.Scrollbar(text_frame, orient=tk.VERTICAL)
        scrollbar_x = tk.Scrollbar(text_frame, orient=tk.HORIZONTAL)
        self.result_text = tk.Text(
            text_frame,
            wrap=tk.NONE,
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set,
            font=("Consolas", 9),
            state=tk.DISABLED,
            bg="#FAFAFA", bd=1, relief=tk.SUNKEN,
            padx=6, pady=4,
        )
        scrollbar_y.config(command=self.result_text.yview)
        scrollbar_x.config(command=self.result_text.xview)

        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.result_text.pack(fill=tk.BOTH, expand=True)

        # ── Patička (tvůrce) ──────────────────────────────────────────────────
        # Pack PŘED status_frame → vizuálně pod status barem (BOTTOM packing)
        footer = tk.Frame(self, bg="#F0F0F0", bd=0)
        footer.pack(fill=tk.X, side=tk.BOTTOM)
        tk.Label(
            footer,
            text="PARAMA Software  \u2502  info@parama.cz",
            font=("Segoe UI", 7), fg="#999999", bg="#F0F0F0",
            anchor=tk.E, padx=8, pady=2,
        ).pack(fill=tk.X)

        # ── Stavový řádek ─────────────────────────────────────────────────────
        status_frame = tk.Frame(self, bd=1, relief=tk.SUNKEN)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        self.status_var = tk.StringVar(value="Vyberte soubor Excel a klikněte na Spustit kontrolu.")
        tk.Label(
            status_frame, textvariable=self.status_var,
            font=("Segoe UI", 8), fg="#555555",
            anchor=tk.W, padx=8, pady=3,
        ).pack(fill=tk.X)

    def _configure_tags(self):
        t = self.result_text
        t.tag_configure("header",       font=("Consolas", 9, "bold"),  foreground="#1A1A1A")
        t.tag_configure("meta",         font=("Consolas", 9),          foreground="#555555")
        t.tag_configure("skipped",      font=("Consolas", 9, "italic"), foreground="#888888")
        t.tag_configure("sheet_header", font=("Consolas", 9, "bold"),  foreground="#0055AA")
        t.tag_configure("sheet_ok",     font=("Consolas", 9),          foreground="#1A7A1A")
        t.tag_configure("row_err",      font=("Consolas", 9, "bold"),  foreground="#CC0000")
        t.tag_configure("cell_err",     font=("Consolas", 9),          foreground="#AA3300")
        t.tag_configure("order_err",    font=("Consolas", 9),          foreground="#C05A00")
        t.tag_configure("summary",      font=("Consolas", 9, "bold"),  foreground="#1A1A1A")
        t.tag_configure("normal",       font=("Consolas", 9),          foreground="#1A1A1A")

    def _pick_file(self):
        path = filedialog.askopenfilename(
            title="Vyberte soubor Excel",
            filetypes=[
                ("Excel soubory", "*.xlsx *.xlsm *.xls"),
                ("Všechny soubory", "*.*"),
            ],
        )
        if path:
            self.path_var.set(path)
            self.run_btn.config(state=tk.NORMAL)
            self.status_var.set("Soubor vybrán. Klikněte na Spustit kontrolu.")

    def _run_validation(self):
        path = self.path_var.get()
        if not path:
            return
        if not os.path.isfile(path):
            messagebox.showerror("Chyba", f"Soubor nebyl nalezen:\n{path}")
            return

        self.run_btn.config(state=tk.DISABLED)
        self.save_btn.config(state=tk.DISABLED)
        self.status_var.set("Probíhá kontrola, čekejte prosím…")
        self.update_idletasks()

        try:
            result = self.validator.validate_file(path)
            self.last_result = result
            self._display_result(result)
            self.save_btn.config(state=tk.NORMAL)

            sheets_with_errors = sum(
                1 for s in result.sheet_results if s.checked and s.row_errors
            )
            order_ok = not result.sheet_order_violations
            if result.total_errors == 0 and order_ok:
                self.status_var.set("Kontrola dokončena – žádné chyby nenalezeny.")
            elif result.total_errors == 0 and not order_ok:
                self.status_var.set(
                    f"Kontrola dokončena – data bez chyb, "
                    f"ale pořadí listů není abecední "
                    f"({len(result.sheet_order_violations)} listů na špatné pozici)."
                )
            elif result.total_errors > 0 and order_ok:
                self.status_var.set(
                    f"Kontrola dokončena – nalezeno {result.total_errors} problematických řádků "
                    f"na {sheets_with_errors} listech."
                )
            else:
                self.status_var.set(
                    f"Kontrola dokončena – nalezeno {result.total_errors} problematických řádků "
                    f"na {sheets_with_errors} listech; "
                    f"pořadí listů není abecední "
                    f"({len(result.sheet_order_violations)} listů na špatné pozici)."
                )
        except Exception as exc:
            self._show_plain_text(f"CHYBA při čtení souboru:\n\n{exc}")
            self.status_var.set("Chyba při čtení souboru – viz výsledky.")
        finally:
            self.run_btn.config(state=tk.NORMAL)

    def _display_result(self, result: ValidationResult):
        segments = self.formatter.segments_for_display(result)
        t = self.result_text
        t.config(state=tk.NORMAL)
        t.delete("1.0", tk.END)
        for text, tag in segments:
            t.insert(tk.END, text, tag)
        t.config(state=tk.DISABLED)
        t.see("1.0")

    def _show_plain_text(self, text: str):
        t = self.result_text
        t.config(state=tk.NORMAL)
        t.delete("1.0", tk.END)
        t.insert(tk.END, text, "row_err")
        t.config(state=tk.DISABLED)

    def _save_report(self):
        if not self.last_result:
            return
        default_name = (
            "report_kapacity_"
            + datetime.now().strftime("%Y%m%d_%H%M%S")
            + ".txt"
        )
        path = filedialog.asksaveasfilename(
            title="Uložit report",
            defaultextension=".txt",
            initialfile=default_name,
            filetypes=[("Textové soubory", "*.txt"), ("Všechny soubory", "*.*")],
        )
        if path:
            plain = self.formatter.format_plain(self.last_result)
            with open(path, "w", encoding="utf-8-sig") as f:
                f.write(plain)
            self.status_var.set(f"Report uložen: {path}")


# ── Spuštění ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = ValidatorApp()
    app.mainloop()
